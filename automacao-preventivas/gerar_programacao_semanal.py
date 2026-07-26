# -*- coding: utf-8 -*-
"""
Gerador automatico da Programacao Semanal de Preventivas - Garagem Quatai (046).
Puxa a tabela ultimo_plano (Athena->Supabase, projeto IMPORTACAO_DADOS) via Supabase CLI,
calcula o que foi FEITO na semana + a PROGRAMACAO da proxima semana, e gera um Excel
no template da oficina (aba 10.000 e aba 5.000, cada uma com "Realizado" em cima e
"Programacao" embaixo).

Roda sozinho (agendado). Nao precisa de argumentos.
"""
import os, sys, json, statistics, datetime, urllib.request

# ---------------- CONFIG ----------------
PROJECT_REF = 'ubppprgquekozluvsloo'          # IMPORTACAO_DADOS
IMPORT_DIR  = r'C:\Users\Guilh\Repositorios\Sistemas\PONTO\importador_supabase'  # credencial Supabase do importador
TELEGRAM_DIR = r'C:\Users\Guilh\Repositorios\PREVENTIVAS\config'  # bot proprio das preventivas (@Preventivas_quatai_bot)
SAIDA_DIR   = r'C:\Users\Guilh\Repositorios\PREVENTIVAS\saidas'
FEITOS_MANUAL_PATH = os.path.join(SAIDA_DIR, 'feitos_manual.json')

def feitos_manual():
    """Carros marcados como feitos manualmente (antes do sistema atualizar). {veic: {codigos_plano}}.
    tipo 10K reseta a Rev.Pesada (2306) E a Inspecao (2305); 5K reseta so a 2305."""
    try:
        d = json.loads(open(FEITOS_MANUAL_PATH, encoding='utf-8').read())
    except Exception:
        return {}
    m = {}
    for it in d.get('itens', []):
        codes = {'2306', '2305'} if it.get('tipo') == '10K' else {'2305'}
        m.setdefault(str(it['veic']), set()).update(codes)
    return m

def feitos_manual_itens():
    """Lista crua dos feitos manuais: [{veic, tipo, data}]."""
    try:
        d = json.loads(open(FEITOS_MANUAL_PATH, encoding='utf-8').read())
        return [dict(veic=str(it['veic']), tipo=it.get('tipo', '10K'), data=it.get('data', ''))
                for it in d.get('itens', [])]
    except Exception:
        return []
EXCLUIR     = {'110797', 'PKB3382'}            # veiculos parados
CONCESS     = {'2645', '2646'}                 # planos de garantia (Euro6)
GAR_FEITOS  = {'242522','242520','242517','242514','242505','242513'}  # garantias ja feitas (marcar OK) - manter manual
WINDOW_KM   = 3000                             # conciliacao de satelites na preventiva 10k
MECS        = ['ANDERSON', 'LUIZ H', 'MAURILIO']

# id_plano -> rotulo curto
LBL = {'2305':'Insp 5.000','2306':'Rev.Pesada 10k','726':'Óleo Motor','1299':'Filtro Ar',
 '2167':'Limpeza Geral','757':'Óleo Câmbio','758':'Óleo Difer.','1300':'Cubos',
 '2314':'Filtro Arla','2965':'Limp.Tq Arla','2966':'Limpeza DPF','2345':'Filtro APU',
 '1132':'Rev.Embreagem','1585':'Fluido Embr.','2505':'Past./Fl.Freio','2309':'Filtro Hidr.',
 '2311':'Serpentina','1239':'Tacógrafo'}
SAT = ['726','1299','2167','757','758','1300','2314','2965','2966','2345','1132','1585','2505','2309','2311']
# quadros de servico (id_plano) na ordem do template
BOXES = [
 ('TROCA DE ÓLEO E FILTROS', ['726']), ('TROCA DE ÓLEO DE CAMBIO', ['757']),
 ('TROCA DE ÓLEO DE DIFERENCIAL', ['758']), ('TROCA DO FILTRO DE AR', ['1299']),
 ('TROCA DO RETENTOR E GRAXA DO CUBO DIANTEIRO', ['1300']),
 ('TROCA FILTRO APU / SERPENTINAS', ['2345','2311']), ('TROCA FILTRO ARLA', ['2314']),
 ('TROCA FILTRO HIDRAULICO', ['2309']), ('', []), ('AFERIÇÃO DE TACÓGRAFO', ['1239']),
 ('REVISÃO NO SISTEMA EMBREAGEM', ['1132']), ('TROCA DE FLUÍDO EMBREAGEM', ['1585']),
 ('LIMPEZA DO TANQUE DE ARLA E PESCADOR', ['2965']), ('LIMPEZA DO FILTRO DPF', ['2966'])]

def log(m): print('[%s] %s' % (datetime.datetime.now().strftime('%H:%M:%S'), m), flush=True)

# ---------------- SUPABASE (REST API, mesma credencial do importador) ----------------
def _sb_config():
    cfg = json.loads(open(os.path.join(IMPORT_DIR, '_supabase.json'), encoding='utf-8').read())
    key = open(os.path.join(IMPORT_DIR, '_supabase_service.key'), encoding='utf-8').read().strip()
    return cfg['url'].rstrip('/'), key

def puxar_dados():
    url, key = _sb_config()
    cols = ('nr_ordem,id_plano,qt_km_intervalo,qt_dia_intervalo,km_rodado,'
            'km_para_proxima,dias_vencido,nr_hodometro,dt_fechamento_os,dt_abertura_os,data_abastecimento')
    headers = {'apikey': key, 'Authorization': 'Bearer ' + key}
    rows = []; step = 1000; off = 0
    while True:
        u = '%s/rest/v1/ultimo_plano?cs_ativo=eq.S&select=%s&limit=%d&offset=%d' % (url, cols, step, off)
        req = urllib.request.Request(u, headers=headers)
        with urllib.request.urlopen(req, timeout=90) as resp:
            batch = json.loads(resp.read().decode('utf-8'))
        rows += batch
        if len(batch) < step: break
        off += step
    return rows

# ---------------- CALCULO ----------------
def num(x):
    try: return float(x)
    except: return None

def montar(rows, hoje):
    cars = {}
    for r in rows:
        v = r['nr_ordem']
        if v in EXCLUIR: continue
        c = cars.setdefault(v, dict(veic=v, kmdias=[], byplan={}, gar=False))
        kr = num(r['km_rodado']); dv = num(r['dias_vencido']); qkm = num(r['qt_km_intervalo'])
        if qkm and qkm > 0 and kr and dv and dv > 30: c['kmdias'].append(kr/dv)
        c['byplan'][r['id_plano']] = r
        if r['id_plano'] in CONCESS: c['gar'] = True
    for c in cars.values():
        c['kmdia'] = round(statistics.median(c['kmdias']), 1) if c['kmdias'] else None

    def kmp(c, code):
        r = c['byplan'].get(code)
        if r is None: return None
        k = num(r['km_para_proxima'])
        return k
    def remkm(c, code):
        k = kmp(c, code); return None if k is None else -k
    def dias(c, code):
        r = c['byplan'].get(code)
        if r is None: return None
        qkm = num(r['qt_km_intervalo']); dv = num(r['dias_vencido']); kd = c['kmdia']
        if qkm == 0: return None if dv is None else round(-dv)
        k = kmp(c, code)
        if k is None or not kd or kd <= 0: return None
        return round(-k/kd)

    def fdate(s):
        s = s or ''
        if len(s) >= 10:
            try: return datetime.date(int(s[:4]), int(s[5:7]), int(s[8:10]))
            except: return None
        return None

    for c in cars.values():
        c['oleo'] = dias(c,'726'); c['insp5'] = dias(c,'2305'); c['rev10'] = dias(c,'2306')

    # ---- FEITAS na semana (ultimos 7 dias) ----
    ini = hoje - datetime.timedelta(days=7)
    feitas_prev, feitas_insp = [], []
    for c in cars.values():
        r10 = c['byplan'].get('2306'); d10 = fdate(r10['dt_fechamento_os']) if r10 else None
        if d10 and ini <= d10 <= hoje: feitas_prev.append(('046-'+c['veic'], d10.strftime('%d/%m')))
        r5 = c['byplan'].get('2305'); d5 = fdate(r5['dt_fechamento_os']) if r5 else None
        if d5 and ini <= d5 <= hoje: feitas_insp.append(('046-'+c['veic'], d5.strftime('%d/%m')))
    feitas_prev.sort(key=lambda x: x[1]); feitas_insp.sort(key=lambda x: x[1])

    # ---- PROGRAMACAO proxima semana ----
    # TODOS os carros fazem 10.000/5.000 in-house (inclusive os de garantia).
    # A garantia so eh "a parte" na revisao 30k/60k da concessionaria (aba Garantia).
    inhouse = list(cars.values())
    def tipo(c):
        r10, i5 = c['rev10'], c['insp5']
        if r10 is not None and (i5 is None or r10 <= i5): return '10K'
        if i5 is not None: return '5K'
        return '10K' if c['oleo'] is not None else '5K'
    def gat10(c):
        vs = [x for x in (c['rev10'], c['oleo']) if x is not None]; return min(vs) if vs else None
    for c in inhouse: c['tipo'] = tipo(c)
    # tira quem foi feito nos ultimos 7 dias (ja aparece em "realizado") + feitos manuais
    fm = feitos_manual()
    # Exclusao por-plano: quem fez a REVISAO 10k recente sai da fila10 (e da fila5, pois 10k reseta a 5k);
    # quem fez SO a INSPECAO 5k sai da fila5, mas CONTINUA na fila10 se a revisao dele estiver vencida.
    prev_set = set(v.replace('046-','') for v,_ in feitas_prev)   # fez a 10.000 recente
    insp_set = set(v.replace('046-','') for v,_ in feitas_insp)   # fez a 5.000 recente
    fila10 = sorted([c for c in inhouse if c['tipo']=='10K' and gat10(c) is not None
                     and c['veic'] not in prev_set and '2306' not in fm.get(c['veic'], set())], key=gat10)
    fila5  = sorted([c for c in inhouse if c['tipo']=='5K'  and c['insp5'] is not None
                     and c['veic'] not in (prev_set | insp_set) and '2305' not in fm.get(c['veic'], set())], key=lambda c: c['insp5'])

    # 10.000 comecam AMANHA e vao ate a SEXTA desta semana (hoje so tem inspecao 5K a noite).
    _d = hoje + datetime.timedelta(days=1)
    fri = hoje + datetime.timedelta(days=(4 - hoje.weekday()))   # sexta desta semana
    if fri < _d: fri = fri + datetime.timedelta(days=7)          # se hoje eh sexta/fds, proxima sexta
    dias_sem = []
    while _d <= fri:
        if _d.weekday() < 5: dias_sem.append(_d)
        _d += datetime.timedelta(days=1)
    _NOMES = ['Segunda-Feira','Terça-Feira','Quarta-Feira','Quinta-Feira','Sexta-Feira','Sábado','Domingo']
    DOW = [_NOMES[d.weekday()] for d in dias_sem]
    ND = len(dias_sem)

    def precisa(c, code):
        if code == '1239':
            d = dias(c, '1239'); return d is not None and d <= 15
        rk = remkm(c, code); return rk is not None and rk <= WINDOW_KM

    carros10 = fila10[:3*ND]
    box_veic = {t: [] for t,_ in BOXES if t}
    esc10 = []; prog_itens = []
    for di in range(ND):
        for c in carros10[di*3:di*3+3]:
            esc10.append(('', '046-'+c['veic'], di))
            prog_itens.append(dict(veic=c['veic'], tipo='10K', data=dias_sem[di].isoformat()))
            for t, codes in BOXES:
                if t and any(precisa(c, cd) for cd in codes): box_veic[t].append('046-'+c['veic'])

    esc5_hoje = [('', '046-'+c['veic']) for c in fila5[:3]]
    for c in fila5[:3]:
        prog_itens.append(dict(veic=c['veic'], tipo='5K', data=hoje.isoformat()))
    esc5_sem = []
    for di, c3 in enumerate([fila5[3+di*3:3+di*3+3] for di in range(ND)]):
        for c in c3:
            esc5_sem.append(('', '046-'+c['veic'], di))
            prog_itens.append(dict(veic=c['veic'], tipo='5K', data=dias_sem[di].isoformat()))

    # ---- GARANTIA (Euro6): proxima revisao da concessionaria pelo hodometro, save 500km ----
    GAR_BUFFER = 500
    gar = []
    for c in cars.values():
        if not c['gar']: continue
        kd = c['kmdia']
        odom = 0
        for r in c['byplan'].values():
            odom = max(odom, (num(r['nr_hodometro']) or 0) + (num(r['km_rodado']) or 0))
        if not kd or kd <= 0 or not odom: continue
        faltam = None
        for code in CONCESS:
            rk = remkm(c, code)
            if rk is None: continue
            if faltam is None or rk < faltam: faltam = rk
        if faltam is None: continue
        done = c['veic'] in GAR_FEITOS
        # ultima vez que o KM do veiculo foi atualizado (abastecimento)
        kmupd = None
        for r in c['byplan'].values():
            d = fdate(r.get('data_abastecimento'))
            if d and (kmupd is None or d > kmupd): kmupd = d
        km_atraso = (hoje - kmupd).days if kmupd else None
        milestone = int(round((odom + faltam) / 30000.0) * 30000)  # grid da concessionaria (30/60/90k)
        vence_d = faltam / kd; alvo_d = (faltam - GAR_BUFFER) / kd
        vence = hoje + datetime.timedelta(days=int(round(vence_d)))
        alvo = hoje + datetime.timedelta(days=max(0, int(round(alvo_d))))
        gar.append(dict(veic='046-'+c['veic'], odom=int(odom), milestone=milestone, falta=int(round(faltam)),
            kmdia=kd, vence=vence.strftime('%d/%m/%Y'), alvo=alvo.strftime('%d/%m/%Y'),
            alvo_sort=alvo.isoformat(), done=done,
            km_upd=kmupd.strftime('%d/%m') if kmupd else '—', km_atraso=km_atraso))
    # pendentes por data de chamada; feitos (OK) no fim
    gar.sort(key=lambda x: (x['done'], x['alvo_sort']))
    # frescor do dado = data mais recente da base (odometro OU OS aberta/fechada)
    km_ref = None; km_leitura = None
    for c in cars.values():
        for r in c['byplan'].values():
            for campo in ('data_abastecimento', 'dt_fechamento_os', 'dt_abertura_os'):
                d = fdate(r.get(campo))
                if d and (km_ref is None or d > km_ref): km_ref = d
            dl = fdate(r.get('data_abastecimento'))
            if dl and (km_leitura is None or dl > km_leitura): km_leitura = dl

    return dict(dias_sem=[d.strftime('%d/%m/%Y') for d in dias_sem], dow=DOW,
        esc10=esc10, boxes=[t for t,_ in BOXES], box_veic=box_veic,
        hoje=hoje.strftime('%d/%m/%Y'), esc5_hoje=esc5_hoje, esc5_sem=esc5_sem,
        feitas_prev=feitas_prev, feitas_insp=feitas_insp, garantia=gar, prog_itens=prog_itens,
        km_ref=km_ref.strftime('%d/%m/%Y') if km_ref else '—',
        km_ref_atraso=(hoje - km_ref).days if km_ref else None,
        km_leitura=km_leitura.strftime('%d/%m/%Y') if km_leitura else '—',
        semana_ini=ini.strftime('%d/%m'), semana_fim=hoje.strftime('%d/%m'))

# ---------------- EXCEL ----------------
def gerar_xlsx(D, caminho):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.properties import PageSetupProperties
    TEALD='1E4A3C'; TEAL='2E6B54'; TAN='EAD9BF'; SLATE='48555E'; ZEB='F3F6F4'; ZEBB='EAF1EC'; NVF='E7F0EB'
    GREENOK='DFF0E4'; OKT='1E7D4F'
    Ar = lambda **k: Font(name='Arial', **k)
    liL = Side(style='thin', color='C8D0D4'); liM = Side(style='thin', color='9AA7AE')
    bIn = Border(liL, liL, liL, liL)
    def fill(h): return PatternFill('solid', fgColor=h)
    CEN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    CENL = Alignment(horizontal='center', vertical='center')
    def Lc(c): return get_column_letter(c)

    def titulo(ws, texto, span, row=1):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        t = ws.cell(row,1,texto); t.font = Ar(bold=True, size=15, color='FFFFFF'); t.fill = fill(TEALD); t.alignment = CEN
        ws.row_dimensions[row].height = 28

    def realizado(ws, row0, titulo_txt, itens, span):
        ws.merge_cells(start_row=row0, start_column=1, end_row=row0, end_column=span)
        h = ws.cell(row0,1, '%s  (%s a %s)' % (titulo_txt, D['semana_ini'], D['semana_fim']))
        h.font = Ar(bold=True, size=11, color='FFFFFF'); h.fill = fill(OKT); h.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row0].height = 20
        r = row0 + 1
        if not itens:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
            ws.cell(r,1,'(nenhuma concluida no periodo / aguardando atualizacao do sistema)').font = Ar(size=9, italic=True, color='6B7787')
            return r + 2
        percol = 6
        cols = min(span, percol*2)
        k = 0
        maxr = r
        for (veic, dia) in itens:
            cc = 1 + (k % 6) * 2
            rr = r + (k // 6)
            v = ws.cell(rr, cc, veic); v.font = Ar(size=9, bold=True, color='1E3A2E'); v.fill = fill(GREENOK); v.alignment = CENL; v.border = bIn
            d = ws.cell(rr, cc+1, dia); d.font = Ar(size=9, color=OKT); d.fill = fill(GREENOK); d.alignment = CENL; d.border = bIn
            maxr = max(maxr, rr); k += 1
        return maxr + 2

    def grade(ws, dias, dow, escala, nv, ndays, r_day):
        r_date, r_sub, r0 = r_day+1, r_day+2, r_day+3
        ws.row_dimensions[r_day].height = 18; ws.row_dimensions[r_date].height = 16; ws.row_dimensions[r_sub].height = 22
        for i in range(ndays):
            c = 1 + i*2
            ws.merge_cells(start_row=r_day, start_column=c, end_row=r_day, end_column=c+1)
            x = ws.cell(r_day, c, dow[i].upper()); x.font = Ar(bold=True, size=10, color='FFFFFF'); x.fill = fill(TEAL); x.alignment = CEN
            ws.merge_cells(start_row=r_date, start_column=c, end_row=r_date, end_column=c+1)
            y = ws.cell(r_date, c, dias[i]); y.font = Ar(bold=True, size=10, color='4A3A20'); y.fill = fill(TAN); y.alignment = CEN
            for lbl, off in [('VEÍCULO',0),('NÍVEL',1)]:
                s = ws.cell(r_sub, c+off, lbl); s.font = Ar(bold=True, size=8.5, color='FFFFFF'); s.fill = fill(SLATE); s.alignment = CEN; s.border = bIn
            for k, rr in enumerate(range(r0, r0+3)):
                ws.row_dimensions[rr].height = 20
                for off in range(2):
                    cell = ws.cell(rr, c+off); cell.border = bIn; cell.alignment = CENL; cell.font = Ar(size=10)
                    if off == 1: cell.fill = fill(NVF)
                    elif k % 2: cell.fill = fill(ZEB)
        from collections import defaultdict
        pd = defaultdict(list)
        for e in escala: pd[e[2]].append(e)
        for di, items in pd.items():
            c = 1 + di*2
            for row_i, it in enumerate(items):
                veic = it[1]
                rr = r0 + row_i
                ws.cell(rr, c, veic).font = Ar(size=10, bold=True, color='1E3A2E')
                ws.cell(rr, c+1, nv).font = Ar(size=9.5, bold=True, color=TEAL)
        return r0 + 3

    def boxes_grid(ws, box_titles, box_veic, start_row):
        box_cols = [(1,2),(3,4),(5,6),(7,8),(9,10)]; VH = 9; PERROW = 5
        for band in range(3):
            hr = start_row + band*(VH+2); ws.row_dimensions[hr].height = 24
            for bi in range(PERROW):
                idx = band*PERROW + bi
                if idx >= len(box_titles): break
                t = box_titles[idx]; c0, c1 = box_cols[bi]
                ws.merge_cells(start_row=hr, start_column=c0, end_row=hr, end_column=c1)
                h = ws.cell(hr, c0, t); h.alignment = CEN
                if t: h.font = Ar(bold=True, size=8.3, color='FFFFFF'); h.fill = fill(TEAL)
                else: h.fill = fill('E9EDEF')
                for cc in range(c0, c1+1): ws.cell(hr, cc).border = Border(liM, liM, liM, liM)
                vs = box_veic.get(t, [])
                for k in range(VH):
                    rr = hr+1+k
                    ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c1)
                    cell = ws.cell(rr, c0, vs[k] if k < len(vs) else None)
                    cell.font = Ar(size=9, bold=bool(k < len(vs)), color='24303A'); cell.alignment = CENL
                    for cc in range(c0, c1+1):
                        ws.cell(rr, cc).border = bIn
                        if k % 2: ws.cell(rr, cc).fill = fill(ZEBB)

    wb = Workbook()
    # ---- ABA 10.000 ----
    ws = wb.active; ws.title = '10.000'
    titulo(ws, 'PROGRAMAÇÃO DE PREVENTIVA SEMANAL  ·  10.000 KM  ·  QUATAÍ', 10)
    r = realizado(ws, 3, 'PREVENTIVAS 10.000 REALIZADAS NA SEMANA', D['feitas_prev'], 10)
    n10 = len(D['dias_sem'])
    r = grade(ws, D['dias_sem'], D['dow'], [tuple(x) for x in D['esc10']], '10.000', n10, r+1)
    ws.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=max(10, n10*2))
    sl = ws.cell(r+1,1,'SERVIÇOS DA SEMANA — conciliados na preventiva (satélite vencendo a ≤%d km)' % WINDOW_KM)
    sl.font = Ar(bold=True, size=9.5, color=TEAL)
    boxes_grid(ws, D['boxes'], D['box_veic'], r+2)
    for c in range(1, max(10, n10*2)+1): ws.column_dimensions[Lc(c)].width = 17 if c%2==1 else 10
    # ---- ABA 5.000 ----
    ws2 = wb.create_sheet('5.000'); ndays = 1 + len(D['dias_sem'])
    titulo(ws2, 'PROGRAMAÇÃO DE INSPEÇÃO SEMANAL  ·  5.000 KM  ·  QUATAÍ', ndays*2)
    r = realizado(ws2, 3, 'INSPEÇÕES 5.000 REALIZADAS NA SEMANA', D['feitas_insp'], ndays*2)
    dias6 = [D['hoje']] + D['dias_sem']; dow6 = ['HOJE'] + D['dow']
    esc5 = [(m, v, 0) for (m, v) in D['esc5_hoje']] + [(e[0], e[1], e[2]+1) for e in D['esc5_sem']]
    r = grade(ws2, dias6, dow6, esc5, '5.000', ndays, r+1)
    ws2.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=ndays*2)
    ws2.cell(r+1,1,'Inspeção 5.000 — apenas inspeção, sem troca de peça.  ·  3 hoje + o restante na semana que vem.').font = Ar(size=9, italic=True, color='6B7787')
    for c in range(1, ndays*2+1): ws2.column_dimensions[Lc(c)].width = 17 if c%2==1 else 10

    # ---- ABA GARANTIA ----
    GARH = '8A4B12'; GARBG = 'FCEBD8'; OKTX = '1E7D4F'
    ws3 = wb.create_sheet('Garantia')
    OKF = 'DFF0E4'
    ws3.merge_cells('A1:H1')
    tt = ws3.cell(1,1,'GARANTIA (EURO6) — REVISÃO NA CONCESSIONÁRIA  ·  QUATAÍ'); tt.font = Ar(bold=True, size=14, color='FFFFFF'); tt.fill = fill(GARH); tt.alignment = CEN
    ws3.row_dimensions[1].height = 28
    ws3.merge_cells('A2:H2')
    ws3.cell(2,1,'Projeção pelo hodômetro. "Chamar em" já desconta o save de 500 km. OK = já realizado. Fora da programação in-house.').font = Ar(size=9, italic=True, color='6B7787')
    hdr = ['VEÍCULO','ODÔMETRO','REVISÃO','FALTA (KM)','KM/DIA','VENCE','CHAMAR EM','STATUS']
    for j, h in enumerate(hdr, 1):
        cc = ws3.cell(4, j, h); cc.font = Ar(bold=True, size=9, color='FFFFFF'); cc.fill = fill(GARH); cc.alignment = CEN; cc.border = bIn
    for i, g in enumerate(D.get('garantia', []), 5):
        done = g.get('done')
        vals = [g['veic'], format(g['odom'], ',d').replace(',', '.'), '%dk' % (g['milestone']//1000),
                format(g['falta'], ',d').replace(',', '.'), g['kmdia'],
                g['vence'], ('—' if done else g['alvo']), ('OK ✓' if done else 'PROGRAMAR')]
        rowfill = OKF if done else GARBG
        for j, v in enumerate(vals, 1):
            cell = ws3.cell(i, j, v)
            col = '1A1D24'
            if j == 7 and not done: col = OKTX
            if j == 8: col = OKTX if done else '8A4B12'
            cell.font = Ar(size=10, bold=(j in (1,7,8)), color=col)
            cell.fill = fill(rowfill); cell.alignment = Alignment(horizontal='left' if j==1 else 'center', vertical='center'); cell.border = bIn
        ws3.row_dimensions[i].height = 18
    for col, w in zip('ABCDEFGH', [12,12,9,11,9,12,12,11]): ws3.column_dimensions[col].width = w
    ws3.sheet_view.showGridLines = False

    for sh in wb.worksheets:
        sh.page_setup.orientation = 'landscape'; sh.page_setup.fitToWidth = 1; sh.page_setup.fitToHeight = 0
        sh.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        sh.page_margins.left = sh.page_margins.right = 0.3; sh.page_margins.top = sh.page_margins.bottom = 0.4
        sh.sheet_view.showGridLines = False
    wb.save(caminho)

# ---------------- MAIN ----------------
def main():
    hoje = datetime.date.today()
    log('Iniciando geracao (%s)' % hoje.strftime('%d/%m/%Y'))
    log('Puxando ultimo_plano do Supabase...')
    rows = puxar_dados()
    log('  %d linhas.' % len(rows))
    D = montar(rows, hoje)
    log('Feitas: %d prev, %d insp | Programado: %d prev, %d+%d insp | Garantia: %d'
        % (len(D['feitas_prev']), len(D['feitas_insp']), len(D['esc10']), len(D['esc5_hoje']), len(D['esc5_sem']), len(D.get('garantia', []))))
    os.makedirs(SAIDA_DIR, exist_ok=True)
    nome = 'Programacao_Semanal_%s.xlsx' % hoje.strftime('%Y-%m-%d')
    caminho = os.path.join(SAIDA_DIR, nome)
    gerar_xlsx(D, caminho)
    log('Gerado: %s' % caminho)
    # ---- TRAVA a programacao: vira o plano oficial da semana (nao muda ate o proximo gerar) ----
    # Guarda a programacao COMPLETA (nao so os itens) pro painel renderizar o plano travado,
    # e o acompanhamento conferir contra ele. Transnet e T-1, entao o plano fica fixo e o
    # "feito" chega nos dias seguintes.
    vig = dict(
        gerado_em=hoje.isoformat(),
        semana_ini=D['dias_sem'][0], semana_fim=D['dias_sem'][-1],
        itens=D.get('prog_itens', []),
        # estrutura completa da programacao (pro painel travado)
        dias_sem=D['dias_sem'], dow=D['dow'], hoje=D['hoje'],
        esc10=D['esc10'], esc5_hoje=D['esc5_hoje'], esc5_sem=D['esc5_sem'],
        box_veic=D['box_veic'], boxes=D['boxes'], km_ref=D.get('km_ref'),
    )
    # ponteiro pro plano vigente
    with open(os.path.join(SAIDA_DIR, 'programacao_vigente.json'), 'w', encoding='utf-8') as f:
        json.dump(vig, f, ensure_ascii=False, indent=1)
    # historico travado (uma copia por geracao, nao sobrescreve as anteriores)
    hist_dir = os.path.join(SAIDA_DIR, 'programacoes_travadas')
    os.makedirs(hist_dir, exist_ok=True)
    with open(os.path.join(hist_dir, 'programacao_%s.json' % hoje.strftime('%Y-%m-%d')), 'w', encoding='utf-8') as f:
        json.dump(vig, f, ensure_ascii=False, indent=1)
    log('Programacao TRAVADA (%d itens) - semana %s a %s.' % (len(vig['itens']), vig['semana_ini'][:5], vig['semana_fim'][:5]))
    return caminho

if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        log('ERRO: %s' % e)
        import traceback; traceback.print_exc()
        sys.exit(1)
